import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_rows(input_excel, output_excel):
    # 读取数据
    df = pd.read_excel(input_excel, engine='openpyxl')

    # 自动识别 idx_r* 列
    idx_cols = [c for c in df.columns if c.startswith('idx_r')]
    if not idx_cols:
        raise ValueError("未找到以 'idx_r' 开头的列，请检查列名。")

    # 条件：NPU_grad > GPU_grad 且 任意 idx_r* > 990
    condition = (df['NPU_grad'] > df['GPU_grad']) & (df[idx_cols] > 990).any(axis=1)

    # 打开工作簿用于着色
    wb = load_workbook(input_excel)
    ws = wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for ridx, row in df[condition].iterrows():
        excel_row = ridx + 2          # pandas索引 + 表头
        step = row.get('Step', ridx)   # 没有 Step 列则用索引代替

        npu = float(row['NPU_grad'])
        gpu = float(row['GPU_grad'])
        if gpu == 0:
            pct_text = "GPU_grad 为 0，无法计算百分比"
            diff_text = f"绝对差值：{npu - gpu:.6g}"
        else:
            inc_pct = (npu - gpu) / gpu * 100
            pct_text = f"NPU_grad 比 GPU_grad 高 {inc_pct:.2f}%"
            diff_text = f"绝对差值：{npu - gpu:.6g}"

        # 找到 >990 的 idx 列与值
        high_pairs = row[idx_cols][row[idx_cols] > 990]

        # —— 控制台输出（中文 + 分隔线 + 空行）——
        print("\n==================================================")
        print(f"命中行：Excel第 {excel_row} 行 | Step = {step}")
        print("--------------------------------------------------")
        print(f"NPU_grad：{npu:.6g}")
        print(f"GPU_grad：{gpu:.6g}")
        print(f"{pct_text}；{diff_text}\n")

        if len(high_pairs) > 0:
            print("下列 idx 值明显偏高（> 990）：")
            for col, val in high_pairs.items():
                print(f"  · 第 {int(step)} 个 step 的 {col} 值为 {val}")
        else:
            print("未发现 >990 的 idx 值（按理不应发生，因为已筛选）")
        print("==================================================\n")

        # 标黄整行
        for c in range(1, len(row) + 1):
            ws.cell(row=excel_row, column=c).fill = yellow

    wb.save(output_excel)


highlight_rows("input_file.xlsx", "output_file.xlsx")
